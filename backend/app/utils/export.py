from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime


def export_orders_to_excel(orders) -> BytesIO:
    """Export orders to Excel with formatted headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "销售记录"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "订单平台", "订单号", "房型", "客人姓名", "销售员", "酒店名称",
        "人数", "预订日期", "状态", "间数", "成本价", "销售价",
        "毛利", "毛利率", "确认号", "备注", "录入时间",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_idx = 2
    for order in orders:
        # One row per order item
        items = order.items if hasattr(order, "items") and order.items else []
        if not items:
            # Write order-level row with empty item fields
            _write_order_row(ws, row_idx, order, None, thin_border)
            row_idx += 1
        else:
            for item in items:
                _write_order_row(ws, row_idx, order, item, thin_border)
                row_idx += 1

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, row_idx)),
            default=10,
        )
        ws.column_dimensions[chr(64 + col)].width = min(max_len + 2, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _write_order_row(ws, row, order, item, border):
    """Write a single data row."""
    data = [
        order.order_platform,
        order.order_number,
        order.room_name,
        order.guest_name,
        order.salesperson,
        order.hotel_name,
        order.guest_count,
        str(order.booking_date),
        order.order_status,
        item.room_count if item else "",
        float(item.cost_price) if item else "",
        float(item.sale_price) if item else "",
        float(item.gross_profit) if item else "",
        f"{float(item.profit_margin)}%" if item else "",
        item.confirmation_number if item else "",
        item.remarks if item else order.other_remarks or "",
        order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
    ]
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = border
        cell.alignment = Alignment(vertical="center")
